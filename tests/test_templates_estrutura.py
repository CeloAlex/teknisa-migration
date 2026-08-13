import random
from io import BytesIO

from httpx import AsyncClient
from openpyxl import Workbook
from openpyxl.utils import column_index_from_string

TEMPLATE_CODIGO = "ESTRUTURA"

LINHA_COM_ENDERECO = {
    "A": 5, "B": "01/01/2020", "C": "", "D": "EST001", "E": "01/01/2020",
    "F": "12345678000199", "G": "", "H": "Empresa Teste LTDA", "I": "2062", "J": "4711302",
    "L": "", "M": "OUTROS", "N": "N", "O": "",
    "P": "OUTROS", "Q": "0055", "R": "SP", "S": "SAO PAULO", "T": "RUA",
    "U": "DAS FLORES", "V": "100", "W": "CENTRO", "X": "01310100", "Y": "",
    "Z": "Fantasia Teste", "AA": "", "AB": "", "AC": "", "AD": "",
}

LINHA_SEM_ENDERECO = {
    "A": 5, "B": "01/02/2020", "C": "", "D": "EST002", "E": "01/02/2020",
    "F": "", "G": "", "H": "Empresa Dois LTDA", "I": "", "J": "",
    "L": "", "M": "", "N": "", "O": "",
    "P": "", "Q": "", "R": "", "S": "", "T": "",
    "U": "", "V": "", "W": "", "X": "", "Y": "",
    "Z": "", "AA": "", "AB": "", "AC": "", "AD": "",
}


def _escrever_linha(ws, linha_num: int, valores: dict) -> None:
    for col_letra, valor in valores.items():
        ws.cell(row=linha_num, column=column_index_from_string(col_letra), value=valor)


def _xlsx_estrutura() -> bytes:
    """Monta um XLSX no mesmo layout do arquivo real (aba Dados, cabeçalho na linha 2, dados
    a partir da linha 3): uma linha com endereço completo (dispara o bloco condicional de
    ENDERECOPARC) e uma linha sem endereço (não dispara)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Dados"
    _escrever_linha(ws, 3, LINHA_COM_ENDERECO)
    _escrever_linha(ws, 4, LINHA_SEM_ENDERECO)
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


async def test_preview_estrutura_duas_linhas_validas(client: AsyncClient) -> None:
    arquivo = _xlsx_estrutura()
    nr_org = random.randint(1_000_000, 9_999_999)

    response = await client.post(
        f"/templates/{TEMPLATE_CODIGO}/preview",
        files={"arquivo": ("estrutura.xlsx", arquivo, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"nr_org": str(nr_org)},
    )

    assert response.status_code == 200
    body = response.json()
    assert body["total_linhas"] == 2
    assert body["validos"] == 2
    assert body["rejeitados"] == 0

    linha1, linha2 = body["linhas"]
    assert linha1["campos"]["_TEM_ENDERECO"] is True
    assert linha1["campos"]["NMPRINCIPALPARC"] == "Empresa Teste LTDA"
    assert linha2["campos"]["_TEM_ENDERECO"] is False
    # razão social preenchida (H) -> primeiro_nao_vazio usa ela, mesmo sem endereço
    assert linha2["campos"]["NMPRINCIPALPARC"] == "Empresa Dois LTDA"


LINHA_TIPO_FORA_DO_DOMINIO = {
    "A": 999, "B": "01/01/2020", "C": "", "D": "EST003", "E": "01/01/2020",
    "F": "12345678000199", "G": "", "H": "Empresa Três LTDA", "I": "2062", "J": "4711302",
    "L": "", "M": "OUTROS", "N": "N", "O": "",
    "P": "OUTROS", "Q": "0055", "R": "SP", "S": "SAO PAULO", "T": "RUA",
    "U": "DAS FLORES", "V": "100", "W": "CENTRO", "X": "01310100", "Y": "",
    "Z": "", "AA": "", "AB": "", "AC": "", "AD": "",
}

LINHA_TIPO_1_SEM_DADOS_CADASTRAIS = {
    "A": 1, "B": "01/01/2020", "C": "", "D": "EST004", "E": "01/01/2020",
    "F": "", "G": "", "H": "Empresa Quatro LTDA", "I": "", "J": "",
    "L": "", "M": "", "N": "N", "O": "",
    "P": "", "Q": "", "R": "", "S": "", "T": "",
    "U": "", "V": "", "W": "", "X": "", "Y": "",
    "Z": "", "AA": "", "AB": "", "AC": "", "AD": "",
}


async def test_preview_estrutura_tipo_fora_do_dominio_gera_alerta_nao_bloqueante(client: AsyncClient) -> None:
    """Reproduz o feedback do piloto NUTRIBEM-TOTAL: tipo de estrutura fora do domínio
    conhecido (o piloto relatou o código 75, hoje já incluído no domínio) deve alertar, não
    impedir a importação."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Dados"
    _escrever_linha(ws, 3, LINHA_TIPO_FORA_DO_DOMINIO)
    buffer = BytesIO()
    wb.save(buffer)
    nr_org = random.randint(1_000_000, 9_999_999)

    response = await client.post(
        f"/templates/{TEMPLATE_CODIGO}/preview",
        files={"arquivo": ("estrutura.xlsx", buffer.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"nr_org": str(nr_org)},
    )

    assert response.status_code == 200
    body = response.json()
    assert body["rejeitados"] == 0
    assert body["alertas"] == 1
    regras = [v["regra"] for v in body["linhas"][0]["validacoes"]]
    assert "fora_do_dominio" in regras


async def test_preview_estrutura_tipo_1_sem_dados_cadastrais_gera_alertas_condicionais(client: AsyncClient) -> None:
    """Reproduz o feedback do piloto NUTRIBEM-TOTAL: ausência de CNPJ/CNAE/natureza
    jurídica/tipo de empresa/endereço/FPAS quando o tipo de estrutura é 1 (empresa) gera
    responsabilidade para as áreas resolverem, mas não bloqueia a importação."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Dados"
    _escrever_linha(ws, 3, LINHA_TIPO_1_SEM_DADOS_CADASTRAIS)
    buffer = BytesIO()
    wb.save(buffer)
    nr_org = random.randint(1_000_000, 9_999_999)

    response = await client.post(
        f"/templates/{TEMPLATE_CODIGO}/preview",
        files={"arquivo": ("estrutura.xlsx", buffer.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"nr_org": str(nr_org)},
    )

    assert response.status_code == 200
    body = response.json()
    assert body["rejeitados"] == 0
    regras = {v["regra"] for v in body["linhas"][0]["validacoes"]}
    campos_com_alerta = {v["campo"] for v in body["linhas"][0]["validacoes"] if v["regra"] == "ausencia_condicional"}
    assert regras == {"ausencia_condicional"}
    assert campos_com_alerta == {"CNPJ", "NATJURIDICA", "CDCNAE", "IDTPEMPRESA", "LOGRADOURO", "FPAS"}


async def test_gerar_script_estrutura_pk_sequencial_e_bloco_condicional(client: AsyncClient) -> None:
    arquivo = _xlsx_estrutura()
    nr_org = random.randint(1_000_000, 9_999_999)

    response = await client.post(
        f"/templates/{TEMPLATE_CODIGO}/gerar-script",
        files={"arquivo": ("estrutura.xlsx", arquivo, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"nr_org": str(nr_org)},
    )

    assert response.status_code == 200
    sql = response.text

    # PK sequencial: duas linhas processadas, cada uma reserva um novo NRPARCNEGOCIO a
    # partir do seed 1738 (Key Resolution Service) — primeira linha recebe 1739.
    assert f"VALUES ( {nr_org}, 1739," in sql
    assert f"VALUES ( {nr_org}, 1740," in sql

    # Bloco principal (PARCNEGOCIO/ESTRUTURAM/ESTRUTURAH) sempre gerado, para as duas linhas.
    assert sql.count("INSERT INTO PARCNEGOCIO (") == 2
    assert sql.count("INSERT INTO ESTRUTURAM (") == 2
    assert sql.count("INSERT INTO ESTRUTURAH (") == 2

    # Bloco condicional de endereço só é gerado para a linha que tem tipo de endereço E
    # logradouro preenchidos (Seção 7.6/26.4) — apenas a primeira linha do fixture.
    assert sql.count("INSERT INTO ENDERECOPARC (") == 1
    assert "SAO PAULO" in sql
    assert "Empresa Dois LTDA" in sql  # segunda linha ainda gera o bloco principal

    assert sql.strip().endswith("COMMIT;")
