from io import BytesIO
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

from app.metadata.schemas import TemplateMetadata

LINHA_PLANILHA = "_linha_planilha"


class ArquivoInvalido(Exception):
    pass


def ler_xlsx(conteudo: bytes, template: TemplateMetadata) -> list[dict[str, Any]]:
    """Adapter de Ingestão XLSX (Anexo A / Anexo F): lê a aba/planilha e a linha inicial de
    dados configuradas no template, devolvendo uma linha por registro mapeada por letra de
    coluna (ex.: "A", "B", "C") — o mesmo contrato tabular que qualquer outro adapter (XML,
    API) deve produzir, para que o restante do motor seja agnóstico de formato de entrada."""
    try:
        workbook = load_workbook(BytesIO(conteudo), data_only=True, read_only=True)
    except Exception as exc:  # noqa: BLE001 — qualquer falha de leitura vira erro de negócio tratável
        raise ArquivoInvalido(f"Não foi possível ler o arquivo XLSX: {exc}") from exc

    sheet_name = template.sheet_name
    planilha = workbook[sheet_name] if sheet_name and sheet_name in workbook.sheetnames else workbook[workbook.sheetnames[0]]

    origens_planilha = sorted({c.origem for c in template.campos if not c.vem_do_contexto})
    colunas_por_origem = {origem: column_index_from_string(origem) for origem in origens_planilha}

    data_start_row = template.data_start_row or 2
    linhas: list[dict[str, Any]] = []
    for linha_celulas in planilha.iter_rows(min_row=data_start_row):
        valores: dict[str, Any] = {
            origem: linha_celulas[idx - 1].value for origem, idx in colunas_por_origem.items()
        }
        if all(v is None or str(v).strip() == "" for v in valores.values()):
            continue
        valores[LINHA_PLANILHA] = linha_celulas[0].row
        linhas.append(valores)
    return linhas
